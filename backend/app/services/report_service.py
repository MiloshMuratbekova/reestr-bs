"""Отчёты по реестру БС: формирование файлов Excel и PDF, история.

Данные берутся из того же итогового реестра, что и остальные страницы —
никакого отдельного расчёта у отчётов нет. Отличается только оформление
и то, что выборка ограничивается сверху: реестр целиком в файл не выгружается.

Шрифты PDF
----------
Контур закрытый, скачать шрифт неоткуда, а стандартные шрифты PDF кириллицу
не содержат. Поэтому берётся Bitstream Vera, входящая в состав reportlab, —
кириллица в ней есть. Если в системе найдётся DejaVu, используется она.
"""

from __future__ import annotations

import json
import re
import time
import uuid
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Sequence

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.algorithms.registry_sql import (
    build_registry_sql,
    build_stats_by_algorithm_sql,
)
from app.core.config import settings
from app.core.logging_config import get_logger
from app.db.clickhouse import clickhouse
from app.models import BsReport
from app.services import algorithm_service, registry_service

logger = get_logger(__name__)


class ReportError(RuntimeError):
    """Отчёт не сформирован."""


@dataclass(frozen=True)
class Column:
    key: str
    title: str
    width: int = 20
    #: Показывать ли колонку в PDF: там места меньше, чем в таблице Excel
    in_pdf: bool = True


@dataclass(frozen=True)
class ReportTemplate:
    key: str
    title: str
    description: str
    columns: List[Column]
    #: Какие параметры показывать в интерфейсе
    parameters: List[str] = field(default_factory=list)


REGISTRY_COLUMNS: List[Column] = [
    Column("taxpayer_iin_bin", "БИН ЮЛ", 16),
    Column("taxpayer_name", "Наименование ЮЛ", 42),
    Column("benefeciary_iin_bin", "ИИН бенефициара", 16),
    Column("benefeciary_name", "Бенефициарный собственник", 38),
    Column("status", "Статус", 28),
    Column("algorithms", "Алгоритмы", 20),
    Column("share_percentage", "Доля, %", 10),
    Column("ball3", "Вероятность, %", 12),
    Column("category", "Категория", 18, in_pdf=False),
    Column("ownership_type", "Тип собственности", 24, in_pdf=False),
    Column("document_info", "Документ", 30, in_pdf=False),
    Column("_actual_date", "Дата актуальности", 16, in_pdf=False),
    Column("ball1", "Балл 1", 8, in_pdf=False),
    Column("ball2", "Балл 2", 8, in_pdf=False),
    Column("dop_info", "Дополнительно", 50, in_pdf=False),
]

ALGORITHM_COLUMNS: List[Column] = [
    Column("algorithm_code", "Алгоритм", 12),
    Column("name", "Название", 38),
    Column("priority", "Балл приоритетности", 12),
    Column("company_count", "Компаний", 12),
    Column("beneficiary_count", "Бенефициаров", 14),
    Column("row_count", "Строк в таблице", 16),
]


TEMPLATES: List[ReportTemplate] = [
    ReportTemplate(
        key="registry",
        title="Реестр БС по всем компаниям",
        description=(
            "Итоговая таблица реестра: пары «компания — бенефициарный собственник» "
            "со статусом, сработавшими алгоритмами и вероятностью."
        ),
        columns=REGISTRY_COLUMNS,
        parameters=["limit"],
    ),
    ReportTemplate(
        key="high_risk",
        title="Компании с высоким уровнем риска",
        description=(
            "Строки реестра, где вероятность выше заданного порога. "
            "По умолчанию порог 70 процентов — та же граница, что и в цветовой шкале."
        ),
        columns=REGISTRY_COLUMNS,
        parameters=["limit", "threshold"],
    ),
    ReportTemplate(
        key="nonresidents",
        title="Бенефициары-нерезиденты",
        description="Строки реестра со статусом, содержащим признак нерезидента.",
        columns=REGISTRY_COLUMNS,
        parameters=["limit"],
    ),
    ReportTemplate(
        key="algorithms",
        title="Статистика по алгоритмам",
        description=(
            "Сколько компаний и бенефициаров выявил каждый алгоритм "
            "и сколько строк в его таблице результата."
        ),
        columns=ALGORITHM_COLUMNS,
        parameters=[],
    ),
]

TEMPLATES_BY_KEY: Dict[str, ReportTemplate] = {t.key: t for t in TEMPLATES}

#: Верхняя граница строк в файле. Реестр целиком не выгружается: файл
#: на миллионы строк не откроется, а память сервера уйдёт целиком.
MAX_REPORT_ROWS = 100_000
DEFAULT_REPORT_ROWS = 5_000


def reports_dir() -> Path:
    """Каталог с файлами отчётов внутри тома DATA_DIR."""
    directory = Path(settings.DATA_DIR)
    if not directory.is_absolute():
        directory = (Path(__file__).resolve().parent.parent.parent / settings.DATA_DIR).resolve()
    path = directory / "reports"
    path.mkdir(parents=True, exist_ok=True)
    return path


def templates_payload() -> List[Dict[str, Any]]:
    return [
        {
            "key": t.key,
            "title": t.title,
            "description": t.description,
            "parameters": t.parameters,
            "columns": [{"key": c.key, "title": c.title} for c in t.columns],
        }
        for t in TEMPLATES
    ]


# ---------------------------------------------------------------------------
# Данные отчёта
# ---------------------------------------------------------------------------
async def _fetch_rows(
    session: AsyncSession, template: ReportTemplate, parameters: Dict[str, Any]
) -> List[Dict[str, Any]]:
    tables = await algorithm_service.active_result_tables(session)
    if not tables:
        raise ReportError(
            "Ни один алгоритм ещё не рассчитан — отчёт формировать не из чего. "
            "Запустите пересчёт реестра."
        )

    if template.key == "algorithms":
        rows = await clickhouse.fetch_all(build_stats_by_algorithm_sql(tables))
        names = {a.code: a.name for a in await algorithm_service.list_algorithms(session)}
        for row in rows:
            row["name"] = names.get(row.get("algorithm_code", ""), "")
        return rows

    limit = int(parameters.get("limit") or DEFAULT_REPORT_ROWS)
    limit = max(1, min(MAX_REPORT_ROWS, limit))

    extra: List[str] = []
    if template.key == "high_risk":
        threshold = float(parameters.get("threshold") or 70)
        threshold = max(0.0, min(100.0, threshold))
        # Значение подставляется числом, а не параметром: оно уже приведено
        # к float и зажато в диапазон, строка сюда попасть не может
        extra.append(f"if(b1.ball1 = 0, 0, round(b2.ball2 / b1.ball1 * 100, 2)) > {threshold}")
    elif template.key == "nonresidents":
        extra.append("pr.status LIKE '%нерезидент%'")

    sql = build_registry_sql(
        tables,
        category_source=registry_service.category_source(),
        extra_conditions=extra or None,
        row_limit=limit,
    )
    return await clickhouse.fetch_all(sql)


def _cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (list, tuple)):
        return ", ".join(str(v) for v in value)
    return str(value)


# ---------------------------------------------------------------------------
# Файлы
# ---------------------------------------------------------------------------
def _write_xlsx(
    path: Path, template: ReportTemplate, rows: Sequence[Dict[str, Any]], subtitle: str
) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Отчёт"

    columns = template.columns
    header_fill = PatternFill("solid", fgColor="1D4ED8")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    sheet.cell(row=1, column=1, value=template.title).font = Font(bold=True, size=14)
    sheet.cell(row=2, column=1, value=subtitle).font = Font(size=9, color="6B7280")

    header_row = 4
    for index, column in enumerate(columns, start=1):
        cell = sheet.cell(row=header_row, column=index, value=column.title)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        sheet.column_dimensions[get_column_letter(index)].width = column.width

    for offset, row in enumerate(rows, start=header_row + 1):
        for index, column in enumerate(columns, start=1):
            value = row.get(column.key)
            if isinstance(value, (int, float)) and not isinstance(value, bool):
                cell = sheet.cell(row=offset, column=index, value=value)
            else:
                cell = sheet.cell(row=offset, column=index, value=_cell(value))
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=column.width > 30)

    sheet.freeze_panes = sheet.cell(row=header_row + 1, column=1)
    sheet.auto_filter.ref = (
        f"A{header_row}:{get_column_letter(len(columns))}{header_row + len(rows)}"
    )
    workbook.save(path)


#: Где искать шрифт для PDF. Сначала Linux — там работает продуктивный контур,
#: DejaVu ставится в образ пакетом fonts-dejavu-core. Windows-пути нужны для
#: запуска на машине разработчика.
#:
#: Bitstream Vera, входящая в состав reportlab, в списке отсутствует
#: СОЗНАТЕЛЬНО: кириллицы в ней нет. Файл с ней собирается без ошибок, но
#: русский текст выходит пустым — это хуже понятного отказа.
_FONT_CANDIDATES = [
    ("DejaVuSans", "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"),
    ("LiberationSans", "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf"),
    ("FreeSans", "/usr/share/fonts/truetype/freefont/FreeSans.ttf"),
    ("NotoSans", "/usr/share/fonts/truetype/noto/NotoSans-Regular.ttf"),
    ("ArialWin", "C:/Windows/Fonts/arial.ttf"),
    ("SegoeUI", "C:/Windows/Fonts/segoeui.ttf"),
    ("Tahoma", "C:/Windows/Fonts/tahoma.ttf"),
]

#: Буквы, по которым проверяется пригодность шрифта
_CYRILLIC_PROBE = "БЖЩЭЮЯабвгдёжзийклмнопрстуфхцчшщъыьэюя"

_font_name: Optional[str] = None


def _pdf_font() -> str:
    """Регистрирует шрифт с кириллицей и возвращает его имя.

    Наличие файла шрифта ещё не значит, что в нём есть русские буквы, поэтому
    каждый кандидат проверяется по таблице глифов. reportlab на отсутствующем
    символе не падает — он рисует пустое место, и отчёт получился бы с пустыми
    строками вместо фамилий.
    """
    global _font_name
    if _font_name:
        return _font_name

    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    checked: List[str] = []
    for name, location in _FONT_CANDIDATES:
        path = Path(location)
        if not path.is_file():
            continue
        checked.append(location)
        try:
            font = TTFont(name, location)
            missing = [
                letter
                for letter in _CYRILLIC_PROBE
                if font.face.charToGlyph.get(ord(letter), 0) == 0
            ]
            if missing:
                logger.warning(
                    "Шрифт %s пропущен: нет кириллических символов (%s)",
                    location,
                    "".join(missing[:10]),
                )
                continue
            pdfmetrics.registerFont(font)
        except Exception as exc:  # noqa: BLE001 — пробуем следующий вариант
            logger.warning("Шрифт %s не зарегистрирован: %s", location, exc)
            continue

        logger.info("PDF печатается шрифтом %s (%s)", name, location)
        _font_name = name
        return name

    raise ReportError(
        "Не найден шрифт с поддержкой кириллицы — PDF сформировать нельзя, "
        "выгрузите отчёт в Excel. "
        + (
            f"Проверены файлы: {', '.join(checked)}. "
            if checked
            else "Ни один из ожидаемых файлов шрифтов не найден. "
        )
        + "В образе шрифт ставится пакетом fonts-dejavu-core."
    )


def _write_pdf(
    path: Path, template: ReportTemplate, rows: Sequence[Dict[str, Any]], subtitle: str
) -> None:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    font = _pdf_font()
    columns = [c for c in template.columns if c.in_pdf]

    title_style = ParagraphStyle("title", fontName=font, fontSize=14, leading=17)
    subtitle_style = ParagraphStyle(
        "subtitle", fontName=font, fontSize=8, leading=11, textColor=colors.HexColor("#6B7280")
    )
    cell_style = ParagraphStyle("cell", fontName=font, fontSize=7, leading=9)
    header_style = ParagraphStyle(
        "header", fontName=font, fontSize=7.5, leading=9.5, textColor=colors.white
    )

    document = SimpleDocTemplate(
        str(path),
        pagesize=landscape(A4),
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
        title=template.title,
    )

    data: List[List[Any]] = [[Paragraph(c.title, header_style) for c in columns]]
    for row in rows:
        data.append([Paragraph(_escape(_cell(row.get(c.key))), cell_style) for c in columns])

    total_width = sum(c.width for c in columns) or 1
    available = document.width
    widths = [available * c.width / total_width for c in columns]

    table = Table(data, colWidths=widths, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1D4ED8")),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D1D5DB")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
                ("TOPPADDING", (0, 0), (-1, -1), 2),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
            ]
        )
    )

    document.build(
        [
            Paragraph(_escape(template.title), title_style),
            Spacer(1, 2 * mm),
            Paragraph(_escape(subtitle), subtitle_style),
            Spacer(1, 4 * mm),
            table,
        ]
    )


def _escape(text: str) -> str:
    """Экранирует разметку абзаца reportlab: данные могут содержать «<» и «&»."""
    return (
        str(text)
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
    )


WRITERS: Dict[str, Callable[..., None]] = {"xlsx": _write_xlsx, "pdf": _write_pdf}


# ---------------------------------------------------------------------------
# Формирование и история
# ---------------------------------------------------------------------------
async def generate(
    session: AsyncSession,
    *,
    template_key: str,
    file_format: str = "xlsx",
    parameters: Optional[Dict[str, Any]] = None,
    created_by: str = "",
) -> Dict[str, Any]:
    """Формирует отчёт, кладёт файл в том и записывает его в историю."""
    template = TEMPLATES_BY_KEY.get(template_key)
    if template is None:
        raise ReportError(f"Неизвестный шаблон отчёта: {template_key}")

    file_format = str(file_format or "xlsx").lower()
    if file_format not in WRITERS:
        raise ReportError("Поддерживаются форматы xlsx и pdf")

    parameters = dict(parameters or {})
    if "limit" in template.parameters:
        parameters["limit"] = clamp_report_rows(parameters.get("limit"))

    started = time.perf_counter()
    rows = await _fetch_rows(session, template, parameters)

    moment = datetime.now(timezone.utc)
    subtitle = (
        f"Сформирован {moment.strftime('%d.%m.%Y %H:%M')} UTC · строк: {len(rows)}"
        f" · система «{settings.APP_NAME}», АФМ РК"
    )
    if created_by:
        subtitle += f" · {created_by}"

    file_name = f"{template.key}-{moment.strftime('%Y%m%d-%H%M%S')}-{uuid.uuid4().hex[:8]}.{file_format}"
    path = reports_dir() / file_name

    try:
        WRITERS[file_format](path, template, rows, subtitle)
    except ReportError:
        raise
    except Exception as exc:  # noqa: BLE001 — наружу отдаём понятный текст
        logger.exception("Файл отчёта %s не записан", template.key)
        raise ReportError(f"Не удалось записать файл отчёта: {exc}") from exc

    duration_ms = int((time.perf_counter() - started) * 1000)
    report = BsReport(
        template=template.key,
        title=template.title,
        file_format=file_format,
        file_name=file_name,
        file_size=path.stat().st_size,
        row_count=len(rows),
        parameters=json.dumps(parameters, ensure_ascii=False),
        status="success",
        created_by=created_by,
        duration_ms=duration_ms,
    )
    session.add(report)
    await session.commit()
    await session.refresh(report)

    logger.info(
        "Отчёт %s (%s) сформирован за %d мс: строк %d, файл %s",
        template.key,
        file_format,
        duration_ms,
        len(rows),
        file_name,
    )
    return report_to_dict(report)


def clamp_report_rows(value: Any) -> int:
    """Число строк отчёта режется на сервере, как и все прочие выборки.

    Здесь действует своя граница MAX_REPORT_ROWS, а не MAX_ROWS_PER_QUERY:
    та настройка ограничивает выдачу на экран, где тысяча строк уже избыточна,
    а выгрузка по смыслу крупнее. Значение из запроса при этом всё равно
    зажимается — параметр приходит от клиента и доверять ему нельзя.
    """
    try:
        number = int(value)
    except (TypeError, ValueError):
        number = DEFAULT_REPORT_ROWS
    return max(1, min(MAX_REPORT_ROWS, number))


def report_to_dict(report: BsReport) -> Dict[str, Any]:
    try:
        parameters = json.loads(report.parameters or "{}")
    except json.JSONDecodeError:
        parameters = {}
    return {
        "id": report.id,
        "template": report.template,
        "title": report.title,
        "file_format": report.file_format,
        "file_name": report.file_name,
        "file_size": report.file_size,
        "row_count": report.row_count,
        "parameters": parameters,
        "status": report.status,
        "error": report.error or "",
        "created_by": report.created_by or "",
        "created_at": report.created_at.isoformat() if report.created_at else "",
        "duration_ms": report.duration_ms,
    }


async def list_reports(session: AsyncSession, *, limit: int = 30) -> List[Dict[str, Any]]:
    rows = (
        await session.execute(
            select(BsReport)
            .order_by(BsReport.created_at.desc())
            .limit(max(1, min(200, int(limit))))
        )
    ).scalars().all()
    return [report_to_dict(report) for report in rows]


_SAFE_NAME = re.compile(r"^[A-Za-z0-9._-]+$")


async def get_report_file(session: AsyncSession, report_id: int) -> tuple[Path, BsReport]:
    """Путь к файлу отчёта. Имя проверяется — путь наружу не собирается."""
    report = (
        await session.execute(select(BsReport).where(BsReport.id == report_id))
    ).scalar_one_or_none()
    if report is None:
        raise ReportError("Отчёт не найден")

    if not report.file_name or not _SAFE_NAME.match(report.file_name):
        raise ReportError("Некорректное имя файла отчёта")

    path = reports_dir() / report.file_name
    if not path.is_file():
        raise ReportError(
            "Файл отчёта отсутствует на диске. Возможно, том с данными был пересоздан — "
            "сформируйте отчёт заново."
        )
    return path, report


async def delete_report(session: AsyncSession, report_id: int) -> None:
    report = (
        await session.execute(select(BsReport).where(BsReport.id == report_id))
    ).scalar_one_or_none()
    if report is None:
        raise ReportError("Отчёт не найден")

    if report.file_name and _SAFE_NAME.match(report.file_name):
        path = reports_dir() / report.file_name
        try:
            path.unlink(missing_ok=True)
        except OSError as exc:
            logger.warning("Файл отчёта %s не удалён: %s", path, exc)

    await session.delete(report)
    await session.commit()
